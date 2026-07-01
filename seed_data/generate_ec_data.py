"""
Generator for Luminary Systems Mock Aruba EdgeConnect Orchestrator seed data.

Reads assets_luminary.xlsx rows where seen_by includes 'aruba'
(all 6 sites) and produces appliances.json consumed by app.py.

All IDs and UUIDs are deterministic (md5-derived).
"""

import hashlib
import json
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
ROOT     = Path(__file__).resolve().parent.parent
RAW      = ROOT / "seed_data" / "raw"
_CENTRAL = ROOT.parent / "luminary-demo-docs" / "master-sheet" / "assets_luminary.xlsx"
_LOCAL   = ROOT / "seed_data" / "source" / "assets_luminary.xlsx"
XLSX     = _CENTRAL if _CENTRAL.exists() else _LOCAL

# ---------------------------------------------------------------------------
# Deterministic helpers
# ---------------------------------------------------------------------------
def _md5(s: str) -> str:
    return hashlib.md5(s.encode()).hexdigest()

def make_uuid(seed: str) -> str:
    h = _md5(seed)
    return f"{h[0:8]}-{h[8:12]}-{h[12:16]}-{h[16:20]}-{h[20:32]}"

def make_portal_object_id(seed: str) -> str:
    return _md5(seed)[:24]

# ---------------------------------------------------------------------------
# Site → networkRole  (hub sites anchor the SD-WAN overlay)
# ---------------------------------------------------------------------------
SITE_ROLE = {
    "San Francisco": "1",   # HQ — hub
    "Amsterdam":     "1",   # EMEA hub
    "New York":      "0",   # spoke
    "London":        "0",   # spoke
    "Singapore":     "0",   # spoke
    "Bangalore":     "0",   # spoke
}

SITE_OCTET = {
    "San Francisco": 11,
    "New York":      12,
    "London":        13,
    "Amsterdam":     14,
    "Singapore":     15,
    "Bangalore":     16,
}

SITE_TIMEZONE = {
    "San Francisco": "America/Los_Angeles",
    "New York":      "America/New_York",
    "London":        "Europe/London",
    "Amsterdam":     "Europe/Amsterdam",
    "Singapore":     "Asia/Singapore",
    "Bangalore":     "Asia/Kolkata",
}

# ---------------------------------------------------------------------------
# Load master sheet
# ---------------------------------------------------------------------------
print(f"Reading: {XLSX}")
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active
headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
H = {h: i for i, h in enumerate(headers)}

sdwan_rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    sb  = row[H["seen_by"]] or ""
    cat = row[H["category"]]
    if cat == "sdwan" and "aruba" in sb.lower():
        sdwan_rows.append({
            "hostname":    row[H["hostname"]],
            "ip":          str(row[H["ip_address"]]),
            "mac":         str(row[H["mac_address"]]),
            "manufacturer":row[H["manufacturer"]],
            "model":       row[H["model"]],
            "serial":      row[H["serial"]],
            "location":    row[H["location"]],
        })
wb.close()

print(f"Total sdwan rows: {len(sdwan_rows)}")
from collections import Counter
locs = Counter(r["location"] for r in sdwan_rows)
for loc, cnt in sorted(locs.items()):
    print(f"  {loc}: {cnt}")

# ---------------------------------------------------------------------------
# Build appliance objects
# ---------------------------------------------------------------------------
appliances = []

for idx, r in enumerate(sdwan_rows):
    hostname = r["hostname"]
    loc      = r["location"]
    seed     = f"ec:appliance:{hostname}"
    h        = _md5(seed)

    ne_pk        = f"{idx}.NE"
    uuid_val     = make_uuid(seed)
    dynamic_uuid = make_uuid(f"dynamic:{seed}")
    portal_id    = make_portal_object_id(f"portal:{seed}")
    appliance_id = 175260 + idx + 1

    # Serial from xlsx (EDGE-XXXXX format) — convert to Orchestrator display format
    serial_raw = r["serial"] or f"EDGE-{idx+1:05d}"

    # MAC from xlsx — already colon-separated uppercase
    mac = r["mac"] if r["mac"] and r["mac"] != "None" else f"84:D4:7E:{h[0:2].upper()}:{h[2:4].upper()}:{h[4:6].upper()}"

    # IP from xlsx
    ip = r["ip"] if r["ip"] and r["ip"].startswith("10.") else f"10.{SITE_OCTET.get(loc,11)}.255.{idx+2}"

    # Model mapping
    model_raw = r["model"] or "EdgeConnect EC-S"
    if "SD-Branch" in model_raw:
        model = "EC-SD-B"
        bandwidth = 100000
    elif "EC-S" in model_raw:
        model = "EC-S"
        bandwidth = 200000
    elif "EC-M" in model_raw:
        model = "EC-M"
        bandwidth = 500000
    else:
        model = "EC-V"
        bandwidth = 200000

    appliance = {
        "id":                  ne_pk,
        "nePk":                ne_pk,
        "uuid":                uuid_val,
        "networkRole":         SITE_ROLE.get(loc, "0"),
        "site":                loc,
        "sitePriority":        0,
        "userName":            "admin",
        "password":            None,
        "groupId":             "2.Network",
        "IP":                  ip,
        "ip":                  ip,
        "webProtocolType":     3,
        "serial":              serial_raw,
        "hasUnsavedChanges":   False,
        "rebootRequired":      False,
        "model":               model,
        "hardwareRevision":    "209003001000 Rev 95994",
        "hostName":            hostname,
        "applianceId":         appliance_id,
        "platform":            "VMware",
        "mode":                "router",
        "bypass":              False,
        "softwareVersion":     "9.3.1.0_95994",
        "startupTime":         None,
        "webProtocol":         "BOTH",
        "systemBandwidth":     bandwidth,
        "state":               1,
        "dynamicUuid":         dynamic_uuid,
        "portalObjectId":      portal_id,
        "discoveredFrom":      2,
        "reachabilityChannel": 2,
        "haPeer":              None,
        "zoneList":            None,
        "interfaceList":       None,
        "tagsList":            None,
        "preconfigStatus":     None,
        "suricataVersion":     "6.0.10",
        "signatureFamily":     "5.x",
    }
    appliances.append(appliance)

# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------
from collections import Counter
print(f"\nAppliances: {len(appliances)}")
by_site  = Counter(a["site"] for a in appliances)
by_model = Counter(a["model"] for a in appliances)
by_role  = Counter("hub" if a["networkRole"]=="1" else "spoke" for a in appliances)
for site, cnt in sorted(by_site.items()):
    print(f"  {site}: {cnt}")
print(f"Models: {dict(by_model)}")
print(f"Roles:  {dict(by_role)}")

# ---------------------------------------------------------------------------
# Write
# ---------------------------------------------------------------------------
RAW.mkdir(parents=True, exist_ok=True)
(RAW / "appliances.json").write_text(json.dumps(appliances, indent=2))
print(f"\nWrote {RAW}/appliances.json")
